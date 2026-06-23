"""
Rate Card Excel → extracted_data-style JSON (MainCosts + CountryZoning + AccessorialCosts2).

Rules (per workbook):
  - Discover a **region prefix** (two capital letters) from worksheet names: the first tab
    whose name begins with ``XX `` (two capitals + a space) and does **not** contain ``ZONE``
    or ``ACCESSORIAL`` defines ``XX`` for the whole workbook. If none, the first ``XX `` tab
    that is not Accessorial (including ZONE) is used so CountryZoning can still match.
  - **MainCosts** (and toolbox header metadata): only sheets whose name starts with
    ``{prefix} `` (same two letters + space), excluding names containing ``ZONE`` or ``ACCESSORIAL``.
  - **Ignored pricing tabs** (per region prefix): CZ/BE and FR skip configured Sending/Receiving
    rate sheets (see ``_IGNORED_PRICING_TABS_BY_REGION``).
  - Sheets whose name **starts with** the same **region prefix** and contains ``ZONE``
    supply ``CountryZoning``; if several match, **only the first** in workbook order is read.
  - From each matching sheet, take the **first** rate table only: ``Net Rates``-style
    blocks use a **Market** row plus a **Zone** row; ``zone_headers`` are
    ``Market + "\\n" + Zone`` per column in **Excel column order** (TB / WW / DOM in any
    order). The **Kg** / **Lbs** header column is detected when present; zone prices start
    in the column immediately to its right, and row weights are read from the Kg column.
  - **Additional** UK/export grids (e.g. ``Weight`` | ``Kg`` | ``Zone 702`` …): when the
    toolbox Market/Zone band is not found, a single header row with ``Zone NNN`` labels and
    weight in the column left of ``Kg`` is parsed before the legacy sliding-window path.
  - Emit one MainCosts section per sheet: ``service_type`` (``Movement\\nService``, e.g.
    ``Receiving Rates\\nUPS Standard Multi`` when Movement/Service labels exist),
    ``cost_category``,
    ``weight_unit``, ``tab_name`` (Excel sheet), ``zone_headers``, ``pricing`` — same
    meaning as ``main.process_main_costs`` (``tab_name`` is ordered before ``zone_headers``).
  - ``CountryZoning`` uses the same layout as ``test.py`` / PDF-style lead-in rows + ``ServiceN`` display strings.
  - ``metadata.client`` is the first non-empty cell in the **first 5 rows** (left-to-right,
    top-to-bottom) of the **first** worksheet whose name starts with ``{region_prefix} ``,
    excluding names containing ``ZONE`` or ``ACCESSORIAL``. When ``--client`` is omitted or
    ``Unknown``, that value is used; explicit ``--client`` overrides.
  - Worksheet named ``Accessorials`` (case-insensitive) → ``AccessorialCosts2`` flat rows.

Dependencies:
  - ``openpyxl`` for ``.xlsx``
  - ``pyxlsb`` for ``.xlsb`` (``pip install pyxlsb``)

Example:
  python conversion-to-json.py --input "input/Rate card/MyCard.xlsb" --output processing/from_excel.json
"""

from __future__ import annotations

import argparse
import json
import re
from collections.abc import Sequence
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

# Excel narrow-column display shows ``#####``; underlying numeric is often a sentinel cap.
UNAVAILABLE_RATE_DISPLAY = "9999999.99"

# Same tokens as main.py — data-row CostName values that must not start a new section.
PACKAGE_ROW_COSTNAME_TOKENS = frozenset({
    "pkg", "cntr", "pallet", "env", "doc", "package", "envelope", "document",
})

# Supported workbook formats (openpyxl does not read .xlsb).
WORKBOOK_SUFFIXES = frozenset({".xlsx", ".xlsb"})

SHEET_MAX_ROW = 1200
SHEET_MAX_COL = 128

# --- Country zoning (dynamic ``{XX}`` prefix + ZONES) + AccessorialCosts2; see test.py parity. ---
COUNTRY_ZONING_SERVICE_NAMES: tuple[str, ...] = (
    "express plus",
    "express",
    "express saver",
    "standard",
    "expedited",
    "express freight",
    "express freight Midday",
) * 2

COUNTRY_ZONING_LEAD_ROWS: tuple[dict[str, str], ...] = (
    {"Service1": "Zonnummer Sända", "Service8": "Zonnummer"},
    {
        "Country": "Country",
        "Code": "Code",
        "CustomerCountry": "Customer 's Country",
        "PostCode1": "Postcode",
        "Service1": "Zonenumber Sending",
        "Service8": "Zonenumber Receiving",
    },
)


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v).strip()
    return str(v).strip()


def _row_trim_right(cells: list[str]) -> list[str]:
    out = list(cells)
    while out and not out[-1].strip():
        out.pop()
    return out


def _is_weight_token(s: str) -> bool:
    t = str(s).strip().lower().replace(",", ".")
    if not t:
        return False
    if re.fullmatch(r"#+", t):
        return False
    t = re.sub(r"\s*kg\s*$", "", t, flags=re.IGNORECASE).strip()
    if not t:
        return False
    try:
        float(t)
        return True
    except ValueError:
        return False


# Shipment weights above this in the left block are treated as stray prices (e.g. Env row).
MAX_REASONABLE_SHIPMENT_WEIGHT_KG = 500.0
# Pkg / freight rows may use high bracket weights; still cap absurd values from mis-aligned columns.
MAX_PKG_BRACKET_WEIGHT_KG = 50_000.0
MAX_PALLET_BRACKET_WEIGHT_KG = 100_000.0


def _max_kg_for_cost_category(low_cc: str) -> float:
    if low_cc in ("env", "doc", "package", "envelope"):
        return MAX_REASONABLE_SHIPMENT_WEIGHT_KG
    if low_cc == "pallet":
        return MAX_PALLET_BRACKET_WEIGHT_KG
    return MAX_PKG_BRACKET_WEIGHT_KG


def _format_weight_display(s: str) -> str:
    """Half-up to 2 decimals; integral weights as ``N.0`` (matches prior JSON style)."""
    t = str(s).strip().replace(",", ".")
    t = re.sub(r"\s*kg\s*$", "", t, flags=re.IGNORECASE).strip()
    if not t or re.fullmatch(r"#+", t):
        return ""
    try:
        d = Decimal(t)
    except Exception:
        return t
    if d == d.to_integral_value():
        return str(int(d)) + ".0"
    q = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(q, ".2f")


def _normalize_weight(s: str) -> str:
    return _format_weight_display(s)


def _is_plausible_shipment_weight(s: str) -> bool:
    """True for typical package weight cells; excludes large numbers that are prices."""
    if not _is_weight_token(s):
        return False
    t = str(s).strip().lower().replace(",", ".")
    t = re.sub(r"\s*kg\s*$", "", t, flags=re.IGNORECASE).strip()
    if re.fullmatch(r"#+", t):
        return False
    try:
        v = float(t)
    except ValueError:
        return False
    return 0 < v <= MAX_REASONABLE_SHIPMENT_WEIGHT_KG


def _toolbox_weight_open_bracket_display(s: str) -> str | None:
    """
    Map ``#####`` or the Excel numeric cap (e.g. 9999999.99) to the published unavailable
    sentinel (same as zone price ``#####``). Returns None for normal weight numbers.
    """
    t = str(s).strip()
    if re.fullmatch(r"#+", t):
        return UNAVAILABLE_RATE_DISPLAY
    if not _is_weight_token(t):
        return None
    raw = t.lower().replace(",", ".")
    raw = re.sub(r"\s*kg\s*$", "", raw, flags=re.IGNORECASE).strip()
    try:
        v = float(raw)
    except ValueError:
        return None
    cap = float(UNAVAILABLE_RATE_DISPLAY.replace(",", "."))
    if abs(v - cap) < 0.01 or v >= cap * 0.999:
        return UNAVAILABLE_RATE_DISPLAY
    return None


def _toolbox_movement_service_line(rows: list[list[str]], start_row: int, end_exclusive: int) -> str:
    """
    Read **Movement** and **Service** label rows above the Market/Zone band (Toolbox layout),
    e.g. ``Receiving Rates`` + ``UPS Express Plus`` → ``Receiving Rates\\nUPS Express Plus``.
    """
    movement = ""
    service = ""
    for r in range(max(0, start_row), max(0, end_exclusive)):
        for i, raw in enumerate(rows[r]):
            s = _cell_str(raw).strip()
            if not s:
                continue
            low = s.lower()
            if low.startswith("movement") or low.startswith("movement:"):
                tail = s.split(":", 1)[1].strip() if ":" in s else ""
                if tail:
                    movement = tail
                else:
                    for j in range(i + 1, len(rows[r])):
                        nxt = _cell_str(rows[r][j]).strip()
                        if nxt and not nxt.lower().startswith("service"):
                            movement = nxt
                            break
            elif "movement" in low and ":" in s:
                movement = s.split(":", 1)[1].strip()
            if low.startswith("service") or low.startswith("service:"):
                tail = s.split(":", 1)[1].strip() if ":" in s else ""
                if tail:
                    service = tail
                else:
                    for j in range(i + 1, len(rows[r])):
                        nxt = _cell_str(rows[r][j]).strip()
                        if nxt:
                            service = nxt
                            break
            elif "service" in low and ":" in s and "movement" not in low:
                service = s.split(":", 1)[1].strip()
    parts = [p for p in (movement.strip(), service.strip()) if p]
    return "\n".join(parts)


def _correct_service_type_for_tab(
    service_type: str, sheet_name: str, region_prefix: str | None
) -> str:
    """
    Fix toolbox Service typos using the worksheet name.

    Some AP tabs store ``UPS … Multi p`` instead of ``UPS … Multi AP`` in the Service cell
    (e.g. tab ``CZ E-Std Multi AP``).
    """
    tab_suffix = _pricing_tab_suffix(sheet_name, region_prefix)
    if not tab_suffix or not tab_suffix.rstrip().upper().endswith("AP"):
        return service_type
    st = (service_type or "").strip()
    if not st:
        return service_type
    if "\n" in st:
        movement, service = st.split("\n", 1)
    else:
        movement, service = "", st
    service = service.strip()
    if service.lower().endswith(" p"):
        service = service[:-2] + " AP"
        return f"{movement}\n{service}".strip("\n") if movement else service
    return service_type


def _toolbox_cell_value_after_label(row: list[str], col_idx: int, label_cell: str) -> str:
    """Value for a label cell: text after ``:`` or next non-empty cell to the right."""
    s = label_cell.strip()
    if ":" in s:
        tail = s.split(":", 1)[1].strip()
        if tail:
            return tail
    for j in range(col_idx + 1, len(row)):
        nxt = _cell_str(row[j]).strip()
        if nxt:
            return nxt
    return ""


def _sheet_header_country_and_rate_currency(rows: list[list[str]]) -> tuple[str, str]:
    """
    Toolbox header rows: ``Country:`` / ``Country`` → country name (metadata **carrier**);
    ``Rate Chart Currency:`` → ``document_currency`` (e.g. SEK).
    """
    country = ""
    currency = ""
    hi = min(len(rows), 80)
    for r in range(hi):
        row = rows[r]
        for i, raw in enumerate(row):
            s = _cell_str(raw).strip()
            if not s:
                continue
            low_sp = s.lower()
            low_ns = re.sub(r"\s+", "", low_sp)
            if low_sp.startswith("country:") or low_ns == "country":
                v = _toolbox_cell_value_after_label(row, i, s)
                if v and not country:
                    country = v
            if low_ns.startswith("ratechartcurrency") or low_sp.startswith("rate chart currency"):
                v = _toolbox_cell_value_after_label(row, i, s)
                if v and not currency:
                    currency = v
    return country, currency


def _sheet_client_from_first_five_rows(rows: list[list[str]]) -> str:
    """
    ``metadata.client``: first non-empty cell when scanning the **first 5 rows**
    left-to-right, top-to-bottom (matches toolbox banner, e.g. ``ASSA ABLOY MC 2018``).
    """
    for r in range(min(5, len(rows))):
        row = rows[r]
        for raw in row:
            s = _cell_str(raw).strip()
            if s:
                return s
    return ""


def _extract_workbook_client_name(
    workbook_path: Path, suffix: str, region_prefix: str | None
) -> str:
    """
    Client from the **first** worksheet matching ``{region_prefix} `` pricing tabs,
    using the first non-empty cell in the first five rows.
    """
    found = ""

    def consume_rows(rows: list[list[str]]) -> None:
        nonlocal found
        if found:
            return
        c = _sheet_client_from_first_five_rows(rows)
        if c:
            found = c

    if suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for name in wb.sheetnames:
                if not sheet_matches_region_pricing_tab(name, region_prefix):
                    continue
                consume_rows(_sheet_to_rows_openpyxl(wb[name]))
                if found:
                    break
        finally:
            wb.close()
    else:
        try:
            from pyxlsb import open_workbook
        except ImportError as e:
            raise SystemExit(
                "Reading .xlsb requires the pyxlsb package. Install with:\n"
                "  pip install pyxlsb"
            ) from e
        with open_workbook(str(workbook_path)) as wb:
            for name in wb.sheets:
                if not sheet_matches_region_pricing_tab(name, region_prefix):
                    continue
                with wb.get_sheet(name) as sheet:
                    consume_rows(_sheet_to_rows_pyxlsb(sheet))
                if found:
                    break
    return found


def _extract_workbook_header_carrier_and_currency(
    workbook_path: Path, suffix: str, region_prefix: str | None
) -> tuple[str, str]:
    """
    Scan **region pricing** sheets for toolbox **Country** and **Rate Chart Currency** rows.
    First non-empty per field across sheets (same card header is usually repeated).
    """
    found: list[str] = ["", ""]

    def consume_rows(rows: list[list[str]]) -> None:
        c, cur = _sheet_header_country_and_rate_currency(rows)
        if c and not found[0]:
            found[0] = c
        if cur and not found[1]:
            found[1] = cur

    if suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            for name in wb.sheetnames:
                if not sheet_matches_region_pricing_tab(name, region_prefix):
                    continue
                consume_rows(_sheet_to_rows_openpyxl(wb[name]))
                if found[0] and found[1]:
                    break
        finally:
            wb.close()
    else:
        try:
            from pyxlsb import open_workbook
        except ImportError as e:
            raise SystemExit(
                "Reading .xlsb requires the pyxlsb package. Install with:\n"
                "  pip install pyxlsb"
            ) from e
        with open_workbook(str(workbook_path)) as wb:
            for name in wb.sheets:
                if not sheet_matches_region_pricing_tab(name, region_prefix):
                    continue
                with wb.get_sheet(name) as sheet:
                    consume_rows(_sheet_to_rows_pyxlsb(sheet))
                if found[0] and found[1]:
                    break
    return found[0], found[1]


def _format_price_cell(s: str) -> str:
    """Half-up to 2 decimals; ``#####`` / hash-only cells → published unavailable sentinel."""
    t = str(s).strip().replace(",", ".")
    if re.fullmatch(r"#+", t):
        return UNAVAILABLE_RATE_DISPLAY
    try:
        d = Decimal(t)
    except Exception:
        return s
    q = d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(q, ".2f")


def _looks_like_rate_type(s: str) -> bool:
    low = s.strip().lower()
    if not low:
        return False
    if "per" in low and (
        "pkg" in low or "shp" in low or "kg" in low or "ship" in low or "pallet" in low
    ):
        return True
    if low in ("per shp", "per kg", "per shipment", "per pkg"):
        return True
    return False


def _is_price_or_placeholder(s: str) -> bool:
    t = str(s).strip().replace(",", ".")
    if not t or t == "-":
        return False
    if re.fullmatch(r"#+", t):
        return True
    try:
        float(t)
        return True
    except ValueError:
        return False


def _looks_like_zone_label(s: str) -> bool:
    u = s.strip().upper()
    if not u:
        return False
    for needle in ("TB", "WW", "DOM", "CZ", "ZONE"):
        if needle in u:
            return True
    if u.isdigit() and len(u) <= 4:
        return True
    if len(u) <= 6 and not u.replace(".", "").isdigit():
        return True
    return False


def _try_parse_data_row(cells: list[str]) -> dict | None:
    """Parse one UPS-style data row; return keys including zone_start index or None."""
    cells = _row_trim_right([_cell_str(c) for c in cells])
    if not cells:
        return None

    weight_idx = None
    lim = min(40, len(cells))
    for j in range(lim):
        if not _is_weight_token(cells[j]):
            continue
        if j + 1 < len(cells):
            nxt = cells[j + 1].strip().lower()
            if nxt in PACKAGE_ROW_COSTNAME_TOKENS or _looks_like_rate_type(cells[j + 1]):
                weight_idx = j
                break
    if weight_idx is None:
        for j in range(lim):
            if _is_weight_token(cells[j]):
                weight_idx = j
                break
    if weight_idx is None:
        return None

    i = weight_idx + 1
    weight_s = _normalize_weight(cells[weight_idx])

    cc = ""
    if i < len(cells) and cells[i].strip().lower() in PACKAGE_ROW_COSTNAME_TOKENS:
        cc = cells[i].strip()
        i += 1

    rt = ""
    if i < len(cells) and _looks_like_rate_type(cells[i]):
        rt = cells[i].strip()
        i += 1

    zone_vals = cells[i:]
    zone_prices: dict[str, str] = {}
    for idx, val in enumerate(zone_vals, start=1):
        if not str(val).strip():
            continue
        if _is_price_or_placeholder(val):
            zone_prices[f"Zone{idx}"] = _format_price_cell(str(val).strip())

    if len(zone_prices) < 2:
        return None

    return {
        "weight": weight_s,
        "cost_category_row": cc,
        "rate_type_row": rt,
        "zone_prices": zone_prices,
        "n_zone_cols": len(zone_vals),
        "zone_start": i,
    }


def _row_is_blank(cells: list[str]) -> bool:
    return not any(c.strip() for c in cells)


def _mostly_zone_labels(slice_: list[str]) -> bool:
    nonempty = [x for x in slice_ if x.strip()]
    if len(nonempty) < 2:
        return False
    hits = sum(1 for x in nonempty if _looks_like_zone_label(x) or _is_price_or_placeholder(x) is False)
    return hits >= max(2, len(nonempty) // 2)


def _sheet_to_rows_openpyxl(ws, max_row: int = SHEET_MAX_ROW, max_col: int = SHEET_MAX_COL) -> list[list[str]]:
    mr = min(ws.max_row or 0, max_row)
    mc = min(ws.max_column or 0, max_col)
    if mr < 1 or mc < 1:
        return []
    rows: list[list[str]] = []
    for r in range(1, mr + 1):
        rows.append([_cell_str(ws.cell(row=r, column=c).value) for c in range(1, mc + 1)])
    return rows


def _sheet_to_rows_pyxlsb(sheet, max_row: int = SHEET_MAX_ROW, max_col: int = SHEET_MAX_COL) -> list[list[str]]:
    """Build row matrix from a pyxlsb worksheet (dense ``sheet.rows()``)."""
    rows: list[list[str]] = []
    for ridx, row in enumerate(sheet.rows()):
        if ridx >= max_row:
            break
        if not row:
            rows.append([])
            continue
        line = [_cell_str(getattr(cell, "v", None)) for cell in row[:max_col]]
        rows.append(line)
    return rows


def _merge_zone_headers(
    upper: list[str] | None,
    lower: list[str],
) -> dict[str, str]:
    n = len(lower)
    headers: dict[str, str] = {}
    for i in range(n):
        parts = []
        if upper and i < len(upper) and upper[i].strip():
            parts.append(upper[i].strip())
        if i < len(lower) and lower[i].strip():
            parts.append(lower[i].strip())
        key = f"Zone{i + 1}"
        if len(parts) > 1:
            headers[key] = "\n".join(parts)
        elif parts:
            headers[key] = parts[0]
        else:
            headers[key] = ""
    return headers


def _legacy_table_head_rows(
    rows: list[list[str]], d: int, parsed: dict
) -> tuple[dict[str, str], int, int]:
    """Zone header dict and row indices for title / cost category (legacy path)."""
    zone_start = parsed["zone_start"]
    n_zones = parsed["n_zone_cols"]
    use_two_zone_header_rows = False
    if d >= 2:
        zslice = rows[d - 2][zone_start : zone_start + n_zones]
        left = rows[d - 2][: max(1, zone_start)]
        left_has_weight = any(_is_weight_token(c) for c in left[:40])
        if not left_has_weight and _mostly_zone_labels(zslice):
            use_two_zone_header_rows = True

    if use_two_zone_header_rows:
        zone_upper = rows[d - 2][zone_start : zone_start + n_zones]
        zone_lower = rows[d - 1][zone_start : zone_start + n_zones]
        title_end = d - 2
        header_row_for_category = d - 1
    else:
        zone_upper = None
        zone_lower = rows[d - 1][zone_start : zone_start + n_zones]
        title_end = d - 1
        header_row_for_category = d - 1
    return _merge_zone_headers(zone_upper, zone_lower), title_end, header_row_for_category


def _legacy_zone_headers_junky(zone_headers: dict[str, str]) -> bool:
    if any(_zone_header_value_is_title_junk(v) for v in zone_headers.values()):
        return True
    nonempty = sum(1 for v in zone_headers.values() if (v or "").strip())
    return nonempty < 4


def _title_lines(rows: list[list[str]], end_exclusive: int) -> str:
    lines = []
    for r in range(0, max(0, end_exclusive)):
        parts = [c.strip() for c in rows[r] if c.strip()]
        if parts:
            lines.append(" ".join(parts))
    return "\n".join(lines)


def _section_cost_category(rows: list[list[str]], header_row_idx: int, zone_start: int) -> str:
    left = _row_trim_right(rows[header_row_idx][:zone_start])
    for cell in reversed(left):
        low = cell.strip().lower()
        if low in ("cntr", "env", "pallet", "doc"):
            return cell.strip()
    return "Cntr"


def _detect_weight_unit(rows: list[list[str]], until: int) -> str:
    blob = "\n".join(" ".join(rows[r]) for r in range(min(until, len(rows))))
    if re.search(r"\blb\b|\blbs\b|\bpound", blob, re.IGNORECASE):
        return "Lb"
    return "Kg"


def _find_net_rates_row(rows: list[list[str]], scan: int = 400) -> int | None:
    for i, row in enumerate(rows[: min(scan, len(rows))]):
        blob = " ".join(_cell_str(c) for c in row).lower()
        if "net rates" in blob:
            return i
    return None


def _is_zone_number_cell(z: str) -> bool:
    t = (z or "").strip()
    if not t:
        return False
    if re.fullmatch(r"#+", t):
        return True
    return bool(re.fullmatch(r"-?\d+", t)) and len(t) <= 6


def _extract_zone_primary_number(z_raw: str) -> str | None:
    """First line of a zone cell (handles ``42`` or ``42\\nCZ`` under Market)."""
    z = (z_raw or "").strip()
    if not z:
        return None
    first = z.split("\n", 1)[0].strip()
    if re.fullmatch(r"#+", first):
        return first
    if re.fullmatch(r"-?\d+", first) and len(first) <= 6:
        return first
    return None


def _market_zone_header_pair(m_raw: str, z_raw: str) -> tuple[str, str] | None:
    """
    If this column belongs to the TB/WW/DOM lane header, return (market, zone_number)
    for building ``Market + "\\n" + zone``.
    """
    m = (m_raw or "").strip()
    z = (z_raw or "").strip()
    if not m and not z:
        return None
    m_main = re.split(r"[/\\]", m)[0].strip()
    if len(m_main) > 36:
        return None
    low = m_main.lower()
    if any(
        x in low
        for x in (
            "rate structure",
            "published rates",
            "effective",
            "lane",
            "mrpp",
            "billing",
            "sections",
        )
    ):
        return None
    mat = re.match(r"^(TB|WW|DOM)(?:\s*[-]?\s*(\d+))?\s*$", m_main, re.I)
    if not mat:
        return None
    code = mat.group(1).upper()
    n_from_m = mat.group(2)
    if n_from_m:
        return code, n_from_m
    z_num = _extract_zone_primary_number(z_raw)
    if z_num is not None:
        return code, z_num
    if _is_zone_number_cell((z or "").strip()):
        return code, z.strip()
    return None


def _band_width_at(rows: list[list[str]], market_r: int, c_start: int, max_col: int) -> int:
    if market_r + 1 >= len(rows):
        return 0
    row_m = rows[market_r]
    row_z = rows[market_r + 1]
    w = 0
    while c_start + w < len(row_m) and c_start + w < max_col:
        m = _cell_str(row_m[c_start + w])
        z = _cell_str(row_z[c_start + w])
        if _market_zone_header_pair(m, z) is None:
            break
        w += 1
    return w


def _lane_label_ok(v: str) -> bool:
    """TB/WW/DOM lane labels; optional third line (e.g. ``CZ42`` from Zone + Lane rows)."""
    s = (v or "").strip()
    if not s:
        return False
    low = s.lower()
    if any(
        x in low
        for x in (
            "rate structure",
            "published rates",
            "effective",
            "mrpp",
            "cntr rate type",
            "billing",
            "sections",
        )
    ):
        return False
    parts = [p.strip() for p in s.replace("\r\n", "\n").split("\n") if p.strip()]
    if not parts:
        return False
    m0 = parts[0]
    mat = re.match(r"^(TB|WW|DOM)(?:\s*[-]?\s*(\d+))?\s*$", m0, re.I)
    if not mat:
        return False
    n_from_m = mat.group(2)
    if n_from_m:
        z0 = parts[1] if len(parts) > 1 else ""
        if not z0:
            return True
        return bool(re.fullmatch(r"\d+|#+", z0)) and len(z0) <= 6
    z0 = parts[1] if len(parts) > 1 else ""
    if not z0 or not re.fullmatch(r"\d+|#+", z0) or len(z0) > 6:
        return False
    for extra in parts[2:]:
        if len(extra) > 14:
            return False
        if not re.match(r"^[A-Za-z0-9#-]+$", extra):
            return False
    return True


def _zone_header_value_is_title_junk(v: str) -> bool:
    """Merged ``Rate Structure`` / ``Published Rates`` blocks mistaken for a lane band."""
    low = (v or "").lower()
    return any(
        x in low
        for x in (
            "rate structure",
            "published rates",
            "effective 22 december",
        )
    )


def _validate_toolbox_band(rows: list[list[str]], market_r: int, c0: int, w: int) -> bool:
    if w < 4 or market_r + 1 >= len(rows):
        return False
    hdrs = _build_zone_headers_from_band(rows, market_r, c0, w)
    if any(_zone_header_value_is_title_junk(v) for v in hdrs.values()):
        return False
    if not _lane_label_ok(hdrs.get("Zone1", "")):
        return False
    ok = sum(1 for i in range(1, w + 1) if _lane_label_ok(hdrs.get(f"Zone{i}", "")))
    return ok >= max(4, int(w * 0.55))


def _validate_toolbox_band_permissive(rows: list[list[str]], market_r: int, c0: int, w: int) -> bool:
    """
    Weaker gate used only when strict validation finds no band (e.g. receiving sheets
    with sparse Lane rows). Still rejects title / published-rates columns.
    """
    if w < 4 or market_r + 1 >= len(rows):
        return False
    hdrs = _build_zone_headers_from_band(rows, market_r, c0, w)
    if any(_zone_header_value_is_title_junk(v) for v in hdrs.values()):
        return False
    if any("published rates" in (v or "").lower() for v in hdrs.values()):
        return False
    if not _lane_label_ok(hdrs.get("Zone1", "")):
        return False
    ok = sum(1 for i in range(1, w + 1) if _lane_label_ok(hdrs.get(f"Zone{i}", "")))
    return ok >= max(3, int(w * 0.35))


def _score_toolbox_band(rows: list[list[str]], market_r: int, c0: int, w: int) -> int:
    """
    Prefer wide, plausible Net Rates bands. Zone order follows Excel (TB, WW, DOM in any
    column order); do not assume DOM is first — that favoured the wrong window vs. the
    Kg-aligned band.
    """
    if w < 2:
        return -10**9
    headers = _build_zone_headers_from_band(rows, market_r, c0, w)
    h1 = headers.get("Zone1") or ""
    h2 = headers.get("Zone2") or ""
    z1, z2 = h1.upper(), h2.upper()
    sc = w * 15 + min(200, w * 8)
    if z1.startswith("TB") and "41" in h1:
        sc += 280
    elif z1.startswith("WW") and "5" in h1.replace("\n", ""):
        sc += 260
    elif z1.startswith("DOM") and re.search(r"DOM\s*\n\s*1\s*$", h1.strip(), re.I):
        sc += 120
    if z1.startswith("TB") and z2.startswith("WW") and "5" in h2.replace("\n", ""):
        sc += 140
    return sc


def _leading_repeated_lane_period(n: int, hdrs: dict[str, str]) -> int:
    """
    If lane headers repeat with period ``p`` (e.g. DOM/99, WW/754 twice each), return ``p``.
    Otherwise return ``n`` (no trim).
    """
    if n < 4:
        return n
    for p in range(2, min(n, 24)):
        if n < p * 2:
            continue
        if all(hdrs.get(f"Zone{i}") == hdrs.get(f"Zone{i + p}") for i in range(1, n - p + 1)):
            return p
    return n


def _validate_net_rates_lane_block(rows: list[list[str]], market_r: int, c0: int, w: int) -> bool:
    """
    True when every column in the band is a valid TB/WW/DOM lane (no title junk).
    Used for **Net Rates** blocks with as few as two columns (DOM / WW only).
    """
    if w < 2 or market_r + 1 >= len(rows):
        return False
    hdrs = _build_zone_headers_from_band(rows, market_r, c0, w)
    if any(_zone_header_value_is_title_junk(v) for v in hdrs.values()):
        return False
    if any("published rates" in (v or "").lower() for v in hdrs.values()):
        return False
    if any("rate structure" in (v or "").lower() for v in hdrs.values()):
        return False
    for i in range(1, w + 1):
        if not _lane_label_ok(hdrs.get(f"Zone{i}", "")):
            return False
    return True


def _columns_have_market_zone_pairs(
    rows: list[list[str]], market_r: int, c0: int, n_chk: int
) -> bool:
    if market_r + 1 >= len(rows) or n_chk < 1:
        return False
    row_m = rows[market_r]
    row_z = rows[market_r + 1]
    for k in range(n_chk):
        if c0 + k >= len(row_m):
            return False
        if _market_zone_header_pair(_cell_str(row_m[c0 + k]), _cell_str(row_z[c0 + k])) is None:
            return False
    return True


def _find_net_rates_first_toolbox_band(
    rows: list[list[str]], max_col: int, net_r: int
) -> tuple[int, int, int, int] | None:
    """
    Prefer the Market/Zone band **on or below** the ``Net Rates`` title row — avoids
    picking repeated ``Market DOM WW`` blocks from Published Rates / Rate Structure.
    """
    cands: list[tuple[int, int, int, int]] = []
    r_hi = min(len(rows) - 2, net_r + 50)
    for r in range(net_r, r_hi + 1):
        lim = min(max_col, len(rows[r]))
        for c in range(0, max(0, lim)):
            w = _band_width_at(rows, r, c, max_col)
            if w < 2:
                continue
            if not _validate_net_rates_lane_block(rows, r, c, w):
                continue
            cands.append((r, r + 1, c, w))
    if not cands:
        return None
    # Duplicate ``Published Rates`` blocks can be one column wider than the real Net Rates
    # grid, or a **narrow tail** (last two WW columns) can score as a valid band. Prefer
    # bands whose width is within one of the maximum, then the **leftmost** start column on
    # that row — this anchors on the true Net Rates table before the duplicate to the right
    # (ENV) and rejects a 2-column tail in favour of the full 7-column block (DOC).
    max_w = max(t[3] for t in cands)
    floor = max(2, max_w - 1)
    tier = [t for t in cands if t[3] >= floor]
    tier.sort(key=lambda t: (t[0], t[2]))
    return tier[0]


def _find_first_toolbox_market_zone_band(rows: list[list[str]], max_col: int) -> tuple[int, int, int, int] | None:
    """
    Best (market_row_idx, zone_row_idx, zone_start_col, zone_count) using header-shape
    scoring. ``Zone1`` is whatever appears first in Excel (TB / WW / DOM in any order).
    """
    net_r = _find_net_rates_row(rows)
    if net_r is not None:
        nr_band = _find_net_rates_first_toolbox_band(rows, max_col, net_r)
        if nr_band is not None:
            return nr_band

    def scan_window(
        r_lo: int, r_hi: int, permissive: bool = False
    ) -> list[tuple[tuple[int, int, int, int], tuple]]:
        out: list[tuple[tuple[int, int, int, int], tuple]] = []
        r_hi = min(r_hi, len(rows) - 2)
        validate = _validate_toolbox_band_permissive if permissive else _validate_toolbox_band
        for r in range(max(0, r_lo), r_hi + 1):
            lim = min(max_col, len(rows[r]))
            for c in range(0, max(0, lim)):
                w = _band_width_at(rows, r, c, max_col)
                if w < 4:
                    continue
                if not validate(rows, r, c, w):
                    continue
                sc = _score_toolbox_band(rows, r, c, w)
                out.append(((r, r + 1, c, w), (sc, c)))
        return out

    cands = scan_window(0, len(rows) - 2, False)
    if net_r is not None:
        r_lo = max(0, net_r - 2)
        r_hi = min(len(rows) - 2, net_r + 35)
        cands.extend(scan_window(r_lo, r_hi, False))

    if not cands:
        cands = scan_window(0, len(rows) - 2, True)
        if net_r is not None:
            r_lo = max(0, net_r - 2)
            r_hi = min(len(rows) - 2, net_r + 35)
            cands.extend(scan_window(r_lo, r_hi, True))

    if not cands:
        return None
    # Highest score, then leftmost column (aligns with Kg column + sheet layout).
    cands.sort(key=lambda item: (-item[1][0], item[0][2]))
    return cands[0][0]


def _find_kg_header_column(
    rows: list[list[str]],
    zone_r: int,
    max_col: int,
    r_low: int | None = None,
    prefer_near_col: int | None = None,
) -> int | None:
    """
    Index of the **Kg** / **Lbs** column in the Net Rates unit row (weights live here;
    zone prices start one column to the right).
    """
    r0 = zone_r
    if r_low is not None and zone_r >= r_low:
        r0 = max(zone_r, r_low)
    r_hi = min(len(rows), max(zone_r + 14, r0 + 8))
    strict: list[tuple[int, int]] = []
    loose: list[tuple[int, int]] = []
    for r in range(r0, r_hi):
        row = rows[r]
        lim = min(len(row), max_col)
        blob = " ".join(_cell_str(row[i]).lower() for i in range(min(28, lim)))
        for c in range(lim):
            cell = _cell_str(row[c]).strip().casefold()
            if cell not in ("kg", "kgs", "lb", "lbs"):
                continue
            loose.append((r, c))
            if "rate" in blob or "cntr" in blob or "package" in blob or "pkg" in blob:
                strict.append((r, c))
    pool = strict or loose
    if not pool:
        return None
    if prefer_near_col is not None:
        pool.sort(key=lambda t: (abs(t[1] - prefer_near_col), t[0], t[1]))
    else:
        pool.sort(key=lambda t: (t[0], t[1]))
    return pool[0][1]


def _lane_row_index(rows: list[list[str]], market_r: int, max_col: int = SHEET_MAX_COL) -> int | None:
    """Row index of the ``Lane`` line under Market/Zone (label may not be in column A)."""
    for delta in range(2, 8):
        ri = market_r + delta
        if ri >= len(rows):
            break
        row = rows[ri]
        lim = min(len(row), max_col)
        for c in range(lim):
            if _cell_str(row[c]).strip().lower() == "lane":
                return ri
    return None


def _build_zone_headers_from_band(rows: list[list[str]], market_r: int, c0: int, n: int) -> dict[str, str]:
    lane_r = _lane_row_index(rows, market_r, SHEET_MAX_COL)
    out: dict[str, str] = {}
    for k in range(n):
        m = _cell_str(rows[market_r][c0 + k])
        z = _cell_str(rows[market_r + 1][c0 + k])
        pair = _market_zone_header_pair(m, z)
        if pair:
            code, num = pair
            base = f"{code}\n{num}"
            z_lines = [ln.strip() for ln in z.replace("\r\n", "\n").split("\n") if ln.strip()]
            if (
                len(z_lines) >= 2
                and z_lines[0] == num
                and re.fullmatch(r"[A-Za-z]+", z_lines[1])
            ):
                suf = z_lines[1] + num
                if not base.endswith(suf):
                    base = f"{base}\n{suf}"
            elif len(z_lines) >= 2 and z_lines[0] == num and z_lines[1] and not z_lines[1].isdigit():
                if not base.endswith(z_lines[1]):
                    base = f"{base}\n{z_lines[1]}"
            if lane_r is not None and lane_r < len(rows):
                lv = _cell_str(rows[lane_r][c0 + k]).strip().replace("\r\n", "\n")
                low_lv = lv.lower()
                if (not lv or low_lv == "lane") and k > 0:
                    m_cur = _cell_str(rows[market_r][c0 + k])
                    z_cur = _cell_str(rows[market_r + 1][c0 + k])
                    m_prev = _cell_str(rows[market_r][c0 + k - 1])
                    z_prev = _cell_str(rows[market_r + 1][c0 + k - 1])
                    if m_cur == m_prev and z_cur == z_prev:
                        lv = _cell_str(rows[lane_r][c0 + k - 1]).strip().replace("\r\n", "\n")
                lparts = [p.strip() for p in lv.split("\n") if p.strip()]
                lv1 = lparts[0] if lparts else ""
                if lv1 and lv1.lower() != "lane":
                    if re.fullmatch(r"[A-Za-z]+\d+", lv1):
                        suf = lv1
                    elif re.fullmatch(r"[A-Za-z]+", lv1):
                        suf = lv1 + num
                    else:
                        suf = lv1
                    if suf and not base.endswith(suf):
                        base = f"{base}\n{suf}"
            out[f"Zone{k + 1}"] = base
        else:
            out[f"Zone{k + 1}"] = "\n".join(p for p in (m.strip(), z.strip()) if p)
    return out


def _left_block_text(cells: list[str], zone_start: int) -> str:
    return " ".join(_cell_str(x) for x in cells[:zone_start]).lower()


def _is_cntr_header_row_no_prices(cells: list[str], zone_start: int, n_zones: int) -> bool:
    left = _left_block_text(cells, zone_start)
    if "cntr" not in left or "kg" not in left:
        return False
    band = [_cell_str(cells[i]) for i in range(zone_start, min(len(cells), zone_start + n_zones))]
    prices = sum(1 for v in band if _is_price_or_placeholder(v))
    return prices < 2


def _band_looks_like_market_restart(slice_: list[str]) -> bool:
    hits = 0
    for x in slice_:
        t = (x or "").strip()
        if re.match(r"^(TB|WW|DOM)(?:\s|[-/]|\d|$)", t, re.I):
            hits += 1
    return hits >= 4


def _zone_prices_echo_lane_numbers(
    zone_prices: dict[str, str],
    zone_headers: dict[str, str],
    n: int,
) -> bool:
    """
    True when zone cells repeat the lane index row (41, 42, 505 under TB/WW headers),
    not currency amounts — a common mis-read of the row directly under the header band.
    """
    hits = 0
    compared = 0
    for i in range(1, n + 1):
        zk = f"Zone{i}"
        if zk not in zone_prices or zk not in zone_headers:
            continue
        h = (zone_headers.get(zk) or "").strip()
        if "\n" not in h:
            continue
        lane_num = h.split("\n", 1)[1].strip()
        if not re.fullmatch(r"\d{1,4}", lane_num):
            continue
        pv = (zone_prices.get(zk) or "").strip()
        try:
            pvf = float(pv.replace(",", "."))
        except ValueError:
            continue
        compared += 1
        try:
            if abs(pvf - float(lane_num)) < 0.001:
                hits += 1
        except ValueError:
            pass
    return compared >= 5 and hits >= compared - 1


def _try_parse_pricing_row_in_band(
    cells: list[str],
    zone_start: int,
    n_zones: int,
    kg_col: int | None = None,
) -> dict | None:
    """Parse one pricing row using a fixed zone column band (Toolbox + legacy grid)."""
    cells = [_cell_str(c) for c in cells]
    if kg_col is not None and zone_start != kg_col + 1:
        kg_col = None
    if zone_start + n_zones > len(cells):
        return None
    band = cells[zone_start : zone_start + n_zones]
    zone_prices: dict[str, str] = {}
    for idx, val in enumerate(band, start=1):
        v = str(val).strip()
        if not v:
            continue
        if _is_price_or_placeholder(v):
            zone_prices[f"Zone{idx}"] = _format_price_cell(v)
    if len(zone_prices) < 2:
        return None
    meta_end = kg_col if kg_col is not None else zone_start
    vals = list(zone_prices.values())
    if len(vals) >= 3 and all(re.match(r"^-?\d+$", v) for v in vals):
        try:
            mx = max(int(v) for v in vals)
        except ValueError:
            mx = 9999
        left_txt = _left_block_text(cells, zone_start)
        has_pkg = any(
            _cell_str(cells[j]).strip().lower() in PACKAGE_ROW_COSTNAME_TOKENS for j in range(meta_end)
        )
        if mx < 200 and "per" not in left_txt and not has_pkg:
            return None

    if kg_col is not None:
        left = cells[:kg_col]
    else:
        left = cells[:zone_start]
    cc = ""
    rt = ""
    cc_idx: int | None = None
    for j, t in enumerate(left):
        ts = t.strip()
        if not ts:
            continue
        low_t = ts.lower()
        if low_t in PACKAGE_ROW_COSTNAME_TOKENS:
            if cc_idx is None:
                cc = ts
                cc_idx = j
            continue
        if _looks_like_rate_type(ts) and not rt:
            rt = ts

    low_cc = (cc or "").strip().lower()
    max_kg = _max_kg_for_cost_category(low_cc) if low_cc else MAX_PKG_BRACKET_WEIGHT_KG

    def _pick_weight_from_span(span: list[str]) -> str:
        for t in span:
            ts = t.strip()
            if not ts:
                continue
            if ts.lower() in PACKAGE_ROW_COSTNAME_TOKENS:
                continue
            if _looks_like_rate_type(ts):
                continue
            ob = _toolbox_weight_open_bracket_display(ts)
            if ob is not None and low_cc in ("env", "doc", "package", "envelope"):
                return ob
            if not _is_weight_token(ts):
                continue
            raw = str(ts).strip().lower().replace(",", ".")
            raw = re.sub(r"\s*kg\s*$", "", raw, flags=re.IGNORECASE).strip()
            try:
                val = float(raw)
            except ValueError:
                continue
            if low_cc in ("env", "doc", "package", "envelope"):
                if not _is_plausible_shipment_weight(ts):
                    continue
            elif val > max_kg:
                continue
            return _format_weight_display(ts)
        return ""

    weight = ""
    if kg_col is not None and kg_col < len(cells):
        ws = cells[kg_col].strip()
        if ws:
            w_open = _toolbox_weight_open_bracket_display(ws)
            if w_open is not None:
                weight = w_open
            elif _is_weight_token(ws):
                raw = str(ws).strip().lower().replace(",", ".")
                raw = re.sub(r"\s*kg\s*$", "", raw, flags=re.IGNORECASE).strip()
                try:
                    val = float(raw)
                except ValueError:
                    val = None
                if val is not None:
                    if low_cc in ("env", "doc", "package", "envelope"):
                        if _is_plausible_shipment_weight(ws):
                            weight = _format_weight_display(ws)
                    elif val <= max_kg:
                        weight = _format_weight_display(ws)
    else:
        weight_spans: list[list[str]] = []
        if cc_idx is not None:
            weight_spans.append(left[:cc_idx])
            tail = left[cc_idx + 1 :]
            ti = 0
            while ti < len(tail) and _looks_like_rate_type(tail[ti].strip()):
                ti += 1
            weight_spans.append(tail[ti:])
        else:
            weight_spans.append(left)
        for span in weight_spans:
            weight = _pick_weight_from_span(span)
            if weight:
                break

    if not cc and not rt and not weight:
        return None
    if low_cc in ("env", "doc", "package", "envelope") and weight:
        try:
            wf = float(str(weight).replace(",", "."))
            cap = float(UNAVAILABLE_RATE_DISPLAY.replace(",", "."))
            if wf > MAX_REASONABLE_SHIPMENT_WEIGHT_KG and abs(wf - cap) > 0.02:
                weight = ""
        except ValueError:
            pass
    return {
        "weight": weight,
        "cost_category_row": cc,
        "rate_type_row": rt,
        "zone_prices": zone_prices,
    }


def _build_service_type_lines(
    rows: list[list[str]], end_exclusive: int, start_row: int = 0
) -> str:
    lines: list[str] = []
    for r in range(max(0, start_row), max(0, end_exclusive)):
        parts = [c.strip() for c in rows[r] if c.strip()]
        if not parts:
            continue
        line = " ".join(parts)
        if len(line) > 450:
            continue
        low = line.lower()
        if low.count("cntr") >= 2 and "rate type" in low and "kg" in low:
            continue
        if line.count("TB-") > 8 or line.count("WW-") > 8:
            continue
        if "rate structure" in low and "published rates" in low:
            continue
        lines.append(line)
    return "\n".join(lines)


def _section_cost_category_toolbox(
    rows: list[list[str]], market_r: int, zone_start: int, kg_col: int | None = None
) -> str:
    col_end = kg_col if kg_col is not None else zone_start
    for r in range(max(0, market_r - 4), min(len(rows), market_r + 2)):
        left = _row_trim_right(rows[r][:col_end])
        for cell in reversed(left):
            low = cell.strip().lower()
            if low in ("cntr", "env", "pallet", "doc"):
                return cell.strip()
    return "Cntr"


def extract_first_table_toolbox_layout(rows: list[list[str]], max_col: int = SHEET_MAX_COL) -> dict | None:
    net_r = _find_net_rates_row(rows)
    band = _find_first_toolbox_market_zone_band(rows, max_col)
    if not band:
        return None
    market_r, zone_r, c0, n = band

    zone_headers = _build_zone_headers_from_band(rows, market_r, c0, n)
    p = _leading_repeated_lane_period(n, zone_headers)
    if p < n:
        n = p
        zone_headers = {f"Zone{i + 1}": zone_headers[f"Zone{i + 1}"] for i in range(n)}

    kg_col = _find_kg_header_column(
        rows,
        zone_r,
        max_col,
        r_low=(net_r if net_r is not None and zone_r >= net_r else None),
        prefer_near_col=c0,
    )
    if kg_col is not None:
        c_snap = kg_col + 1
        n_snap = _band_width_at(rows, market_r, c_snap, max_col)
        n_chk = min(n_snap, max(n, 6))
        col_ok = _columns_have_market_zone_pairs(rows, market_r, c_snap, n_chk)
        lane_ok_snap = n_snap >= 2 and _validate_net_rates_lane_block(rows, market_r, c_snap, n_snap)
        wide_ok_snap = n_snap >= 4 and (
            _validate_toolbox_band(rows, market_r, c_snap, n_snap)
            or _validate_toolbox_band_permissive(rows, market_r, c_snap, n_snap)
        )
        if col_ok and (lane_ok_snap or wide_ok_snap):
            c0, n = c_snap, n_snap
            zone_headers = _build_zone_headers_from_band(rows, market_r, c0, n)
            p2 = _leading_repeated_lane_period(n, zone_headers)
            if p2 < n:
                n = p2
                zone_headers = {f"Zone{i + 1}": zone_headers[f"Zone{i + 1}"] for i in range(n)}
        else:
            kg_col = None

    st_start = net_r if net_r is not None else 0
    built = _build_service_type_lines(rows, market_r, start_row=st_start).strip()
    # Movement/Service live above the ``Net Rates`` row; ``st_start`` is ``net_r`` so do not
    # use it as the lower bound for label scanning (would skip Country/Movement/Service rows).
    banner_lo = max(0, market_r - 60)
    mv = _toolbox_movement_service_line(rows, banner_lo, market_r).strip()
    # ``built`` spans Net Rates → Market and picks up banner rows (``Net Rates``, ``Rate Structure``);
    # the canonical product line is Movement + Service only — do not prefix ``built`` when ``mv`` exists.
    if mv:
        service_type = mv
    else:
        service_type = " ".join(built.split()).strip()
    if not service_type.strip():
        service_type = "Receiving Rates"
    cost_category = _section_cost_category_toolbox(rows, market_r, c0, kg_col)
    weight_unit = _detect_weight_unit(rows, market_r)

    pricing: list[dict] = []
    seen_price_row = False
    for r in range(zone_r + 1, len(rows)):
        if _row_is_blank(rows[r]):
            if seen_price_row:
                break
            continue
        row_cells = rows[r]
        if _band_looks_like_market_restart(row_cells[c0 : c0 + n]) and seen_price_row:
            break
        blob = " ".join(_cell_str(x) for x in row_cells).lower()
        if "published rates" in blob and seen_price_row:
            break
        if _is_cntr_header_row_no_prices(row_cells, c0, n):
            continue
        pr = _try_parse_pricing_row_in_band(row_cells, c0, n, kg_col)
        if not pr:
            continue
        if _zone_prices_echo_lane_numbers(pr["zone_prices"], zone_headers, n):
            continue
        zp = {}
        for k, v in pr["zone_prices"].items():
            m = re.match(r"^Zone(\d+)$", k)
            if m and int(m.group(1)) <= n:
                zp[k] = v
        if len(zp) < 2:
            continue
        row_obj: dict = {"weight": pr["weight"], "zone_prices": zp}
        if pr["cost_category_row"]:
            row_obj["cost_category"] = pr["cost_category_row"]
        if pr["rate_type_row"]:
            row_obj["rate_type"] = pr["rate_type_row"]
        pricing.append(row_obj)
        seen_price_row = True

    if not pricing:
        return None

    return {
        "service_type": service_type,
        "cost_category": cost_category,
        "weight_unit": weight_unit,
        "zone_headers": zone_headers,
        "pricing": pricing,
    }


_ZONE_NUMBER_HEADER_RE = re.compile(r"^zone\s+(\d+)\s*$", re.I)

# Product rows above numeric weight brackets (GB export Net Rates); not package tokens on data rows.
_NET_RATES_PRODUCT_ROW_LABELS = frozenset(
    {"envelope", "document", "doc", "env", "package"}
)


def _is_zone_number_header_cell(s: str) -> bool:
    return bool(_ZONE_NUMBER_HEADER_RE.match((s or "").strip()))


def _zone_number_header_width(row: list[str], c0: int, max_col: int) -> int:
    w = 0
    while c0 + w < len(row) and c0 + w < max_col:
        if not _is_zone_number_header_cell(_cell_str(row[c0 + w])):
            break
        w += 1
    return w


def _find_net_rates_zone_number_header_band(
    rows: list[list[str]], max_col: int = SHEET_MAX_COL
) -> tuple[int, int, int, int, int] | None:
    """
    Locate ``Weight`` | (optional unit col) | ``Zone 702`` … header band (GB export layout).

    Returns ``(header_row, zone_start, n_zones, weight_col, unit_col)`` or None.
    """
    net_r = _find_net_rates_row(rows)
    r_lo = max(0, net_r) if net_r is not None else 0
    r_hi = min(len(rows), r_lo + 80)

    for r in range(r_lo, r_hi):
        row = rows[r]
        lim = min(len(row), max_col)
        for c in range(lim):
            if _cell_str(row[c]).strip().casefold() != "weight":
                continue
            for zstart in (c + 1, c + 2):
                if zstart >= lim:
                    continue
                if zstart > c + 1 and not _cell_str(row[zstart - 1]).strip():
                    pass
                w = _zone_number_header_width(row, zstart, max_col)
                if w < 4:
                    continue
                weight_col = c
                unit_col = zstart - 1
                if unit_col <= weight_col:
                    unit_col = weight_col + 1
                return r, zstart, w, weight_col, unit_col
    return None


def _build_zone_headers_from_zone_number_row(
    row: list[str], zone_start: int, n: int
) -> dict[str, str]:
    out: dict[str, str] = {}
    for k in range(n):
        v = _cell_str(row[zone_start + k]).strip()
        out[f"Zone{k + 1}"] = v
    return out


def _try_parse_net_rates_weight_kg_row(
    cells: list[str],
    zone_start: int,
    n_zones: int,
    weight_col: int,
    unit_col: int,
) -> dict | None:
    """
    Parse one pricing row for the Weight | Kg | Zone NNN grid (weight in ``weight_col``,
    unit label in ``unit_col``, prices from ``zone_start``).
    """
    cells = [_cell_str(c) for c in cells]
    if zone_start + n_zones > len(cells):
        return None

    w_raw = cells[weight_col].strip() if weight_col < len(cells) else ""
    u_raw = cells[unit_col].strip().casefold() if unit_col < len(cells) else ""
    low_w = w_raw.lower()

    if not w_raw:
        return None
    if low_w in _NET_RATES_PRODUCT_ROW_LABELS:
        return None
    if low_w in PACKAGE_ROW_COSTNAME_TOKENS and not _is_weight_token(w_raw):
        return None

    open_bracket = _toolbox_weight_open_bracket_display(w_raw)
    if open_bracket is not None:
        weight_out = open_bracket
    elif not _is_weight_token(w_raw):
        return None
    else:
        weight_out = _format_weight_display(w_raw)
        if u_raw in ("kg", "kgs", "lb", "lbs") and not re.search(
            r"\b(?:kg|kgs|lb|lbs)\b", weight_out, re.I
        ):
            weight_out = f"{weight_out} {u_raw}"

    band = cells[zone_start : zone_start + n_zones]
    zone_prices: dict[str, str] = {}
    for idx, val in enumerate(band, start=1):
        v = str(val).strip()
        if not v:
            continue
        if _is_price_or_placeholder(v):
            zone_prices[f"Zone{idx}"] = _format_price_cell(v)
    if len(zone_prices) < 2:
        return None

    return {"weight": weight_out, "zone_prices": zone_prices}


def extract_first_table_net_rates_zone_grid(rows: list[list[str]]) -> dict | None:
    """
    **Additional** extractor for export Net Rates sheets: one header row
    ``Weight`` | ``Kg`` | ``Zone 702`` | ``Zone 703`` | … and numeric weights in the
    weight column (not the legacy “first number in row” heuristic).
    """
    band = _find_net_rates_zone_number_header_band(rows)
    if not band:
        return None

    header_r, zone_start, n, weight_col, unit_col = band
    zone_headers = _build_zone_headers_from_zone_number_row(rows[header_r], zone_start, n)

    net_r = _find_net_rates_row(rows)
    st_start = net_r if net_r is not None else 0
    banner_lo = max(0, header_r - 60)
    mv = _toolbox_movement_service_line(rows, banner_lo, header_r).strip()
    built = _build_service_type_lines(rows, header_r, start_row=st_start).strip()
    if mv:
        service_type = mv
    elif built:
        service_type = built
    else:
        service_type = "Receiving Rates"
    if not service_type.strip():
        service_type = "Receiving Rates"

    cost_category = _section_cost_category(rows, header_r, zone_start)
    weight_unit = _detect_weight_unit(rows, header_r)

    pricing: list[dict] = []
    for r in range(header_r + 1, len(rows)):
        if _row_is_blank(rows[r]):
            # Separator blank between Documents / Packages blocks — keep scanning.
            continue
        row_cells = rows[r]
        blob = " ".join(_cell_str(x) for x in row_cells).lower()
        if pricing and _cell_str(row_cells[weight_col]).strip().casefold() == "weight":
            break
        if pricing and "published rates" in blob:
            break
        pr = _try_parse_net_rates_weight_kg_row(
            rows[r], zone_start, n, weight_col, unit_col
        )
        if not pr:
            continue
        pricing.append({"weight": pr["weight"], "zone_prices": pr["zone_prices"]})

    if not pricing:
        return None

    return {
        "service_type": service_type,
        "cost_category": cost_category,
        "weight_unit": weight_unit,
        "zone_headers": zone_headers,
        "pricing": pricing,
    }


def extract_first_main_costs_table_legacy(rows: list[list[str]]) -> dict | None:
    """
    Find the first contiguous UPS-style grid and return one MainCosts section dict,
    or None if no table was found.
    """
    net_r = _find_net_rates_row(rows)
    start_ri = net_r if net_r is not None else 0
    first_data = None
    d = None
    zone_headers: dict[str, str] = {}
    title_end = 0
    header_row_for_category = 0
    for ri in range(start_ri, len(rows)):
        parsed = _try_parse_data_row(rows[ri])
        if not parsed:
            continue
        zh, te, hcat = _legacy_table_head_rows(rows, ri, parsed)
        if _legacy_zone_headers_junky(zh):
            continue
        first_data = parsed
        d = ri
        zone_headers = zh
        title_end = te
        header_row_for_category = hcat
        break
    if first_data is None or d is None:
        return None

    zone_start = first_data["zone_start"]
    n_zones = first_data["n_zone_cols"]

    title_block = _title_lines(rows, title_end).strip()
    mv = _toolbox_movement_service_line(rows, max(0, d - 60), d).strip()
    if mv:
        service_type = mv
    else:
        service_type = title_block
    if not service_type.strip():
        service_type = "Receiving Rates"

    cost_category = _section_cost_category(rows, header_row_for_category, zone_start)
    weight_unit = _detect_weight_unit(rows, title_end)

    pricing: list[dict] = []
    r = d
    while r < len(rows):
        if _row_is_blank(rows[r]):
            break
        pr = _try_parse_data_row(rows[r])
        if not pr:
            break
        if pr["n_zone_cols"] != n_zones:
            break
        row_obj: dict = {
            "weight": pr["weight"],
            "zone_prices": pr["zone_prices"],
        }
        if pr["cost_category_row"]:
            row_obj["cost_category"] = pr["cost_category_row"]
        if pr["rate_type_row"]:
            row_obj["rate_type"] = pr["rate_type_row"]
        pricing.append(row_obj)
        r += 1

    if not pricing:
        return None

    return {
        "service_type": service_type,
        "cost_category": cost_category,
        "weight_unit": weight_unit,
        "zone_headers": zone_headers,
        "pricing": pricing,
    }


def extract_first_main_costs_table(rows: list[list[str]]) -> dict | None:
    tb = extract_first_table_toolbox_layout(rows)
    if tb:
        return tb
    zone_grid = extract_first_table_net_rates_zone_grid(rows)
    if zone_grid:
        return zone_grid
    return extract_first_main_costs_table_legacy(rows)


# ----- CountryZoning + AccessorialCosts2 (region prefix + ZONES tab; Accessorials). -----


def discover_workbook_region_prefix(sheet_names: Sequence[str]) -> str | None:
    """
    Two-letter region code: first sheet name matching ``^[A-Z]{2}\\s``.

    Prefer a tab without ``ZONE`` / ``ACCESSORIAL`` so pricing sheets define the prefix.
    If none, use the first ``XX `` tab that is not Accessorial (e.g. only ``SE ZONES`` exists).
    """
    first_non_accessorial: str | None = None
    for raw in sheet_names:
        s = (raw or "").strip()
        m = re.match(r"^([A-Z]{2})\s", s)
        if not m:
            continue
        px = m.group(1)
        u = s.upper()
        if "ACCESSORIAL" in u:
            continue
        if first_non_accessorial is None:
            first_non_accessorial = px
        if "ZONE" not in u:
            return px
    return first_non_accessorial


def sheet_matches_region_pricing_tab(name: str, region_prefix: str | None) -> bool:
    """Pricing / toolbox header sheets: ``{prefix} `` at start; no ``ZONE`` / ``ACCESSORIAL``."""
    if not region_prefix:
        return False
    u = (name or "").strip().upper()
    if "ZONE" in u or "ACCESSORIAL" in u:
        return False
    p = region_prefix.strip().upper()
    return len(p) == 2 and u.startswith(p + " ")


def _norm_pricing_tab_suffix(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _pricing_tab_suffix(sheet_name: str, region_prefix: str | None) -> str | None:
    """Part of sheet name after ``{prefix} `` (e.g. ``E-Express Plus_ENV``)."""
    if not region_prefix:
        return None
    p = region_prefix.strip().upper()
    if len(p) != 2:
        return None
    raw = (sheet_name or "").strip()
    prefix = p + " "
    if not raw.upper().startswith(prefix):
        return None
    return _norm_pricing_tab_suffix(raw[len(prefix) :])


def _ignored_pricing_tabs(*tab_suffixes: str) -> frozenset[str]:
    return frozenset(_norm_pricing_tab_suffix(t) for t in tab_suffixes)


# Czech Republic + Belgium: skip these pricing tabs during MainCosts extraction.
_IGNORED_PRICING_TABS_CZ_BE = _ignored_pricing_tabs(
    # Sending Rates
    "E-Express Plus_ENV",
    "E-Express Plus AP_ENV",
    "E-Express Plus_DOC",
    "E-Express Plus AP_DOC",
    "E-Express Plus_PKG",
    "E-Express Plus AP_PKG",
    "E-Express_ENV",
    "E-Express AP_ENV",
    "E-Express_DOC1",
    "E-Express AP_DOC",
    "E-Express_PKG1",
    "E-Express AP_PKG",
    "E-WWE DDP",
    "E-WWE DDU",
    "E-APE Single",
    "E-APE Multi",
    "E-Expedited1",
    "E-Expedited AP",
    "E-WWEF",
    "E-WWEF Mday",
    # Receiving Rates
    "I-Express Plus_ENV",
    "I-Express Plus AP_ENV",
    "I-Express Plus_DOC",
    "I-Express Plus AP_DOC",
    "I-Express Plus_PKG",
    "I-Express Plus AP_PKG",
    "I-Express_ENV",
    "I-Express AP_ENV",
    "I-Express_DOC1",
    "I-Express AP_DOC",
    "I-Express_PKG1",
    "I-Express AP_PKG",
    "I-WWE DDP",
    "I-APE Single",
    "I-APE Multi",
    "I-Expedited1",
    "I-Expedited AP",
    "I-WWEF",
    "I-WWEF Mday",
)

# France: skip these pricing tabs during MainCosts extraction.
_IGNORED_PRICING_TABS_FR = _ignored_pricing_tabs(
    # Sending Rates
    "E-Express Plus_ENV",
    "E-Express Plus AP_ENV",
    "E-Express Plus_DOC",
    "E-Express Plus AP_DOC",
    "E-Express Plus_PKG",
    "E-Express Plus AP_PKG",
    "E-Express AP_ENV",
    "E-Express AP_DOC",
    "E-Express AP_PKG",
    "E-Express Svr AP_ENV",
    "E-Express Svr AP_DOC",
    "E-Express Svr AP_PKG",
    "E-WWE DDP",
    "E-WWE DDU",
    "E-Std Single AP",
    "E-Std Multi AP",
    "E-APE Single",
    "E-APE Multi",
    "E-Expedited",
    "E-Expedited AP",
    "E-WWEF",
    "E-WWEF ToD",
    # Receiving Rates
    "I-Express Plus_ENV",
    "I-Express Plus AP_ENV",
    "I-Express Plus_DOC",
    "I-Express Plus AP_DOC",
    "I-Express Plus_PKG",
    "I-Express Plus AP_PKG",
    "I-Express AP_ENV",
    "I-Express AP_DOC",
    "I-Express AP_PKG",
    "I-Express Svr AP_ENV",
    "I-Express Svr AP_DOC",
    "I-Express Svr AP_PKG",
    "I-WWE DDP",
    "I-WWE DDU",
    "I-Std Single AP1",
    "I-Std Single AP2",
    "I-Std Multi AP1",
    "I-Std Multi AP2",
    "I-APE Single",
    "I-APE Multi",
    "I-Expedited",
    "I-Expedited AP",
    "I-WWEF",
    "I-WWEF ToD",
)

_IGNORED_PRICING_TABS_BY_REGION: dict[str, frozenset[str]] = {
    "CZ": _IGNORED_PRICING_TABS_CZ_BE,
    "BE": _IGNORED_PRICING_TABS_CZ_BE,
    "FR": _IGNORED_PRICING_TABS_FR,
}


def sheet_is_ignored_pricing_tab(sheet_name: str, region_prefix: str | None) -> bool:
    """True when this region's configured ignore list contains the sheet's tab suffix."""
    suffix = _pricing_tab_suffix(sheet_name, region_prefix)
    if suffix is None:
        return False
    ignored = _IGNORED_PRICING_TABS_BY_REGION.get((region_prefix or "").strip().upper())
    if not ignored:
        return False
    return suffix in ignored


def sheet_matches_region_zones_tab(name: str, region_prefix: str | None) -> bool:
    """Zoning sheet: name starts with **region prefix** (after strip) and contains ``ZONE``."""
    if not region_prefix:
        return False
    u = (name or "").strip().upper()
    p = region_prefix.strip().upper()
    return len(p) == 2 and u.startswith(p) and "ZONE" in u


def _workbook_sheet_names(workbook_path: Path, suffix: str) -> list[str]:
    """Worksheet names in workbook order."""
    if suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    if suffix == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except ImportError as e:
            raise SystemExit(
                "Reading .xlsb requires the pyxlsb package. Install with:\n"
                "  pip install pyxlsb"
            ) from e
        with open_workbook(str(workbook_path)) as wb:
            return list(wb.sheets)
    return []


def sheet_matches_accessorials_tab(sheet_name: str) -> bool:
    return sheet_name.strip().lower() == "accessorials"


def _norm_accessorials2_hdr(cell: str) -> str:
    t = _cell_str(cell).strip().lower()
    return re.sub(r"\s+", " ", t)


def _accessorials2_header_cells_match(h: list[str]) -> bool:
    """Seven header cells: Accessorial, Movement, Market, Service, Rate Type, Rate, Section Nbr."""
    if len(h) < 7:
        return False
    if "accessorial" not in h[0]:
        return False
    if h[1] != "movement":
        return False
    if h[2] != "market":
        return False
    if h[3] != "service":
        return False
    if "rate" not in h[4] or "type" not in h[4]:
        return False
    if h[5] != "rate":
        return False
    if "section" not in h[6] or ("nbr" not in h[6] and "number" not in h[6]):
        return False
    return True


def _find_accessorials2_header(rows: list[list[str]]) -> tuple[int, int] | None:
    """Return ``(header_row_index, first_column_index)`` for the Accessorials table, or ``None``."""
    lim = min(len(rows), 120)
    for ri in range(lim):
        row = rows[ri]
        if len(row) < 7:
            continue
        maxj = min(len(row), SHEET_MAX_COL) - 7
        for j in range(0, max(0, maxj + 1)):
            h = [_norm_accessorials2_hdr(row[j + k]) for k in range(7)]
            if _accessorials2_header_cells_match(h):
                return ri, j
    return None


def extract_accessorial_costs2_from_sheet_rows(rows: list[list[str]]) -> list[dict[str, str]]:
    """
    Parse the **Accessorials** tab: blue header row with Accessorial, Movement, Market,
    Service, Rate Type, Rate, Section Nbr — then one JSON object per data row.
    """
    found = _find_accessorials2_header(rows)
    if not found:
        return []
    hdr_i, j0 = found
    out: list[dict[str, str]] = []
    empty_run = 0
    for ri in range(hdr_i + 1, min(len(rows), SHEET_MAX_ROW)):
        row = rows[ri]

        def cell(off: int) -> str:
            c = j0 + off
            return _cell_str(row[c]) if c < len(row) else ""

        cost = cell(0).strip()
        if not cost:
            empty_run += 1
            if empty_run >= 25 and out:
                break
            continue
        empty_run = 0
        if _norm_accessorials2_hdr(cost) == "accessorial":
            continue

        rec: dict[str, str] = {
            "CostName": cost,
            "Movement": cell(1).strip(),
            "Market": cell(2).strip(),
            "Service": cell(3).strip(),
            "RateType": cell(4).strip(),
            "Rate": cell(5).strip(),
            "Section Nbr": cell(6).strip(),
        }
        if not (rec["Movement"] or rec["Market"] or rec["Service"]):
            continue
        out.append(rec)
    return out


def _collect_accessorials2_from_openpyxl(workbook_path: Path) -> list[dict]:
    import openpyxl

    merged: list[dict] = []
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_matches_accessorials_tab(sheet_name):
                continue
            rows = _sheet_to_rows_openpyxl(wb[sheet_name])
            part = extract_accessorial_costs2_from_sheet_rows(rows)
            merged.extend(part)
            if not part and _find_accessorials2_header(rows) is None:
                print(
                    f"[WARN] Sheet {sheet_name!r}: Accessorials 7-column header not found (skipped)"
                )
    finally:
        wb.close()
    return merged


def _collect_accessorials2_from_pyxlsb(workbook_path: Path) -> list[dict]:
    try:
        from pyxlsb import open_workbook
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsb requires the pyxlsb package. Install with:\n"
            "  pip install pyxlsb"
        ) from e

    merged: list[dict] = []
    with open_workbook(str(workbook_path)) as wb:
        for sheet_name in wb.sheets:
            if not sheet_matches_accessorials_tab(sheet_name):
                continue
            with wb.get_sheet(sheet_name) as sheet:
                rows = _sheet_to_rows_pyxlsb(sheet)
            part = extract_accessorial_costs2_from_sheet_rows(rows)
            merged.extend(part)
            if not part and _find_accessorials2_header(rows) is None:
                print(
                    f"[WARN] Sheet {sheet_name!r}: Accessorials 7-column header not found (skipped)"
                )
    return merged


def _zones_norm_header(cell: str) -> str:
    return _cell_str(cell).strip().lower().replace("\r\n", "\n").replace("\n", " ")


def _zones_row_service_match_score(header_row: list[str], c0: int) -> int:
    row = [_cell_str(x) for x in header_row]
    score = 0
    for k in range(14):
        if c0 + k >= len(row):
            break
        blob = _zones_norm_header(row[c0 + k])
        needle = COUNTRY_ZONING_SERVICE_NAMES[k]
        if needle in blob:
            score += 1
            continue
        if needle.replace(" ", "") in blob.replace(" ", ""):
            score += 1
    return score


def _find_postcode_block_column(header: list[str]) -> int | None:
    """
    Index of the **Postcode** header cell. The next column to the right is the second
    postcode column; **Service1** starts two columns to the right of the Postcode header.
    """
    for i, raw in enumerate(header):
        u = _cell_str(raw).strip().lower().replace("'", "").replace("`", "")
        u_nospace = re.sub(r"\s+", "", u)
        if "express" in u or "standard" in u or "expedited" in u or "freight" in u:
            continue
        if "postcode" in u_nospace:
            return i
        if u_nospace.startswith("postco"):
            return i
        if "post" in u and "code" in u_nospace:
            return i
    return None


def _merged_vertical_header_blob(rows: list[list[str]], hdr_i: int, c: int) -> str:
    """Non-empty header text above/below ``hdr_i`` for column ``c`` (merged vertical labels)."""
    parts: list[str] = []
    for r in range(max(0, hdr_i - 12), min(len(rows), hdr_i + 6)):
        if c < len(rows[r]):
            raw = _cell_str(rows[r][c]).strip()
            if raw:
                parts.append(raw.lower().replace("\r\n", " ").replace("\n", " "))
    return " ".join(parts)


def _column_header_matches_service_slot(blob: str, slot: int) -> bool:
    """Header text matches Sending/Receiving slot ``slot`` (0..13); slot 7 repeats *express plus* like slot 0."""
    pos = slot % 7
    if pos == 0:
        return "express plus" in blob
    if pos == 1:
        return (
            "express" in blob
            and "express plus" not in blob
            and "express saver" not in blob
            and "saver" not in blob
        )
    if pos == 2:
        return "express saver" in blob or ("express" in blob and "saver" in blob)
    if pos == 3:
        return "standard" in blob
    if pos == 4:
        return "expedited" in blob
    if pos == 5:
        return "freight" in blob and "midday" not in blob
    if pos == 6:
        return "midday" in blob
    return False


def _collect_service_columns_skip_empty(
    rows: list[list[str]],
    hdr_i: int,
    row_len: int,
    first_c: int,
) -> list[int] | None:
    """
    Build ``Service1``..``Service14`` column indices by scanning **left to right** from
    ``first_c``, **skipping** columns with no service header (gaps / filters between Sending
    and Receiving). Each non-skipped column must match the next expected service in order.
    """
    out: list[int] = []
    want = 0
    for c in range(first_c, min(row_len, first_c + 100)):
        if want >= 14:
            break
        blob = _merged_vertical_header_blob(rows, hdr_i, c).strip()
        if not blob:
            continue
        if _column_header_matches_service_slot(blob, want):
            out.append(c)
            want += 1
    if want != 14:
        return None
    return out


def _find_country_zoning_layout(rows: list[list[str]]) -> dict | None:
    """
    SE_ZONES layout (explicit columns):

    - ``CustomerCountry`` ← header contains *customer* + *country* (Customer's Country).
    - ``Code`` ← cell text exactly ``Code``.
    - ``Country`` ← header exactly ``Country`` (English; column to the right of *Land* when present).
    - ``PostCode1`` / ``PostCode2`` ← column whose header is Postcode (or ``Postco`` …),
      plus the column immediately to its right.
    - ``Service1`` … ``Service14`` ← the next **14 service columns** after the postcode block,
      in sheet order (Sending seven, then Receiving seven). **Blank / headerless columns**
      between Sending and Receiving are **not** counted as services and are skipped.
    """
    hi = min(len(rows), 120)
    hdr_i: int | None = None
    for ri in range(hi):
        row = rows[ri]
        cells = [_cell_str(c).strip() for c in row]
        lows = [c.lower() for c in cells]
        if "code" not in lows:
            continue
        if not any("country" in x for x in lows):
            continue
        hdr_i = ri
        break
    if hdr_i is None:
        return None

    header = [_cell_str(c) for c in rows[hdr_i]]
    c_code: int | None = None
    c_customer: int | None = None
    c_land: int | None = None
    c_country: int | None = None
    post_cols: list[int] = []
    exact_country_cols: list[int] = []
    fuzzy_country_cols: list[int] = []

    for i, raw in enumerate(header):
        s = raw.strip()
        low = s.lower()
        if low == "code":
            c_code = i
        elif "customer" in low and "country" in low:
            c_customer = i
        elif low == "land":
            c_land = i
        elif low == "country":
            exact_country_cols.append(i)
        elif low.startswith("country") and "customer" not in low:
            fuzzy_country_cols.append(i)
        elif "post" in low:
            post_cols.append(i)

    if exact_country_cols:
        if c_land is not None:
            after_land = [j for j in exact_country_cols if j > c_land]
            c_country = after_land[0] if after_land else exact_country_cols[-1]
        else:
            c_country = exact_country_cols[-1]
    elif fuzzy_country_cols:
        if c_land is not None:
            after_land = [j for j in fuzzy_country_cols if j > c_land]
            c_country = after_land[0] if after_land else fuzzy_country_cols[-1]
        else:
            c_country = fuzzy_country_cols[-1]

    if c_code is None:
        return None

    row_len = max(
        (len(rows[j]) for j in range(hdr_i, min(hdr_i + 40, len(rows)))),
        default=len(header),
    )

    post_data_cols: list[int] | None = None
    svc0: int | None = None

    c_post = _find_postcode_block_column(header)
    if c_post is not None:
        post_data_cols = [c_post, c_post + 1]
        svc0 = c_post + 2
    elif c_country is not None:
        # Fallback: two columns after English Country, then services
        svc_try = c_country + 3
        post_data_cols = [c_country + 1, c_country + 2]
        if svc_try < row_len:
            svc0 = svc_try
        else:
            svc_one = c_country + 2
            if svc_one < row_len:
                svc0 = svc_one
                post_data_cols = [c_country + 1]

    if svc0 is None:
        idents = [x for x in (c_customer, c_code, c_land, c_country) if x is not None]
        if post_cols:
            idents.extend(post_cols)
        if not idents:
            return None
        last_ident = max(idents)
        best_c0: int | None = None
        best_sc = -1
        for delta in (0, -1, 1, -2, 2, -3, 3):
            rh = hdr_i + delta
            if rh < 0 or rh >= len(rows):
                continue
            scan = [_cell_str(c) for c in rows[rh]]
            for c0 in range(last_ident + 1, min(row_len - 13, SHEET_MAX_COL - 14)):
                sc = _zones_row_service_match_score(scan, c0)
                if sc > best_sc:
                    best_sc = sc
                    best_c0 = c0
        if best_c0 is None or best_sc < 4:
            best_c0 = last_ident + 1
        svc0 = best_c0

    if svc0 is None:
        return None

    first_after_post: int
    if post_data_cols:
        first_after_post = post_data_cols[-1] + 1
    else:
        first_after_post = svc0

    service_cols = _collect_service_columns_skip_empty(
        rows, hdr_i, row_len, first_after_post
    )
    if not service_cols and svc0 + 13 < row_len:
        service_cols = list(range(svc0, svc0 + 14))
    if not service_cols or len(service_cols) != 14 or any(c >= row_len for c in service_cols):
        return None

    svc0 = service_cols[0]

    return {
        "hdr_i": hdr_i,
        "c_customer": c_customer,
        "c_code": c_code,
        "c_land": c_land,
        "c_country": c_country,
        "post_cols": sorted(set(post_cols)),
        "svc0": svc0,
        "service_cols": service_cols,
        "post_data_cols": post_data_cols,
        "row_len": row_len,
    }


def _country_zoning_service_cell_display(raw_cell: str, service_name: str) -> str:
    """
    Always returns a display string: ``"<service>\\n<zone>"`` when the cell is a whole
    number; otherwise the canonical **service name** (empty / junk cells → service name).
    """
    t = _cell_str(raw_cell).strip()
    if not t or t in ("-", "–", "—", "N/A", "n/a"):
        return service_name
    low = t.lower()
    if low in ("x", "#n/a", "#ref!"):
        return service_name
    if "\n" in t:
        return t.replace("\r\n", "\n")
    t_num = t.replace(",", ".")
    try:
        d = float(t_num)
        if abs(d) < 1e12 and d == int(d):
            return f"{service_name}\n{int(d)}"
    except (ValueError, OverflowError):
        pass
    if re.fullmatch(r"\d+", t):
        return f"{service_name}\n{t}"
    if low == service_name:
        return service_name
    return f"{service_name}\n{t}"


def _is_probable_zoning_data_row(layout: dict, row_cells: list[str]) -> bool:
    cc = layout["c_code"]
    if cc is None or cc >= len(row_cells):
        return False
    code = _cell_str(row_cells[cc]).strip().upper()
    return len(code) == 2 and code.isalpha()


def extract_se_zones_country_zoning(rows: list[list[str]]) -> list[dict]:
    """
    Parse SE_ZONES into ``CountryZoning`` rows (``assa3`` lead-in rows, then data).

    Column rules: **Customer's Country** → ``CustomerCountry``; **Code** → ``Code``;
    header exactly **Country** → ``Country``; **Postcode** column + column to its right
    → ``PostCode1`` / ``PostCode2``; the next 14 **service** columns (skipping blank
    columns between Sending and Receiving) → ``Service1`` … ``Service14``. Every
    ``ServiceN`` is emitted; numeric zone cells use
    ``"<service>\\n<zone>"``, empty cells use the service label only.
    """
    layout = _find_country_zoning_layout(rows)
    if not layout:
        return []
    hdr_i = layout["hdr_i"]
    c_customer = layout["c_customer"]
    c_code = layout["c_code"]
    c_country = layout["c_country"]
    c_land = layout["c_land"]
    svc0 = layout["svc0"]
    service_cols: list[int] = layout.get("service_cols") or list(range(svc0, svc0 + 14))
    fixed_posts = layout.get("post_data_cols")

    if fixed_posts is not None:
        post_data_cols = list(fixed_posts)
    else:
        post_cols = layout["post_cols"]
        hdr_row = rows[hdr_i]
        if post_cols:
            lo = min(post_cols)
            post_data_cols = [c for c in range(lo, svc0)]
        elif c_country is not None:
            post_data_cols = [c for c in range(c_country + 1, svc0) if c < len(hdr_row)]
        elif c_land is not None:
            post_data_cols = [c for c in range(c_land + 1, svc0) if c < len(hdr_row)]
        else:
            post_data_cols = [c for c in range(c_code + 1, svc0) if c < len(hdr_row)]

    out: list[dict] = []
    for ri in range(hdr_i + 1, min(len(rows), SHEET_MAX_ROW)):
        row_cells = [_cell_str(c) for c in rows[ri]]
        if not any(x.strip() for x in row_cells):
            continue
        if not _is_probable_zoning_data_row(layout, row_cells):
            continue

        obj: dict[str, str] = {}
        if c_country is not None and c_country < len(row_cells):
            obj["Country"] = row_cells[c_country].strip()
        elif c_land is not None and c_land < len(row_cells):
            obj["Country"] = row_cells[c_land].strip()
        if c_code is not None and c_code < len(row_cells):
            obj["Code"] = row_cells[c_code].strip().upper()
        if c_customer is not None and c_customer < len(row_cells):
            obj["CustomerCountry"] = row_cells[c_customer].strip()

        if post_data_cols and len(post_data_cols) >= 2:
            for pi in range(2):
                ci = post_data_cols[pi]
                pv = row_cells[ci].strip() if ci < len(row_cells) else ""
                obj[f"PostCode{pi + 1}"] = pv
        else:
            for pi, ci in enumerate((post_data_cols or [])[:6]):
                if ci >= len(row_cells):
                    continue
                pv = row_cells[ci].strip()
                if pv:
                    obj[f"PostCode{pi + 1}"] = pv

        for si in range(14):
            name = COUNTRY_ZONING_SERVICE_NAMES[si]
            ci = service_cols[si]
            raw = row_cells[ci] if ci < len(row_cells) else ""
            obj[f"Service{si + 1}"] = _country_zoning_service_cell_display(raw, name)

        if obj.get("Code"):
            out.append(obj)
    return [*COUNTRY_ZONING_LEAD_ROWS, *out]


def _collect_country_zoning_from_openpyxl(
    workbook_path: Path, region_prefix: str | None
) -> list[dict]:
    if not region_prefix:
        return []
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        target: str | None = None
        for sheet_name in wb.sheetnames:
            if sheet_matches_region_zones_tab(sheet_name, region_prefix):
                target = sheet_name
                break
        if not target:
            return []
        rows = _sheet_to_rows_openpyxl(wb[target])
        z = extract_se_zones_country_zoning(rows)
        if len(z) > len(COUNTRY_ZONING_LEAD_ROWS):
            return z
        print(
            f"[WARN] Sheet {target!r} matched {region_prefix}+ZONE but no country rows parsed"
        )
        return []
    finally:
        wb.close()


def _collect_country_zoning_from_pyxlsb(
    workbook_path: Path, region_prefix: str | None
) -> list[dict]:
    if not region_prefix:
        return []
    try:
        from pyxlsb import open_workbook
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsb requires the pyxlsb package. Install with:\n"
            "  pip install pyxlsb"
        ) from e

    with open_workbook(str(workbook_path)) as wb:
        target: str | None = None
        for sheet_name in wb.sheets:
            if sheet_matches_region_zones_tab(sheet_name, region_prefix):
                target = sheet_name
                break
        if not target:
            return []
        with wb.get_sheet(target) as sheet:
            rows = _sheet_to_rows_pyxlsb(sheet)
        z = extract_se_zones_country_zoning(rows)
        if len(z) > len(COUNTRY_ZONING_LEAD_ROWS):
            return z
        print(
            f"[WARN] Sheet {target!r} matched {region_prefix}+ZONE but no country rows parsed"
        )
    return []


def _country_zoning_and_accessorials2_for_suffix(
    workbook_path: Path, suffix: str, region_prefix: str | None
) -> tuple[list[dict], list[dict]]:
    """Dispatch zoning + accessorial collectors by workbook type (additive helper)."""
    if suffix == ".xlsb":
        return (
            _collect_country_zoning_from_pyxlsb(workbook_path, region_prefix),
            _collect_accessorials2_from_pyxlsb(workbook_path),
        )
    if suffix == ".xlsx":
        return (
            _collect_country_zoning_from_openpyxl(workbook_path, region_prefix),
            _collect_accessorials2_from_openpyxl(workbook_path),
        )
    return [], []


def resolve_rate_card_dir(project_root: Path) -> Path | None:
    for folder in ("Rate card", "Rate Card", "rate card"):
        p = project_root / "input" / folder
        if p.is_dir():
            return p
    return None


def _glob_workbooks(d: Path) -> list[Path]:
    if not d.is_dir():
        return []
    found = list(d.glob("*.xlsx")) + list(d.glob("*.xlsb"))
    return sorted({p.resolve(): p for p in found}.values(), key=lambda p: p.name.lower())


def iter_default_workbook_candidates(project_root: Path):
    """
    Yield .xlsx / .xlsb in priority order (first existing file wins in default_input_workbook):
      1) input/Rate card/ (any recognised spelling)
      2) input/*.xlsx and input/*.xlsb
      3) input/<subfolder>/* for each direct subfolder (skipping duplicate of #1)
    """
    rate_dir = resolve_rate_card_dir(project_root)
    if rate_dir:
        for p in _glob_workbooks(rate_dir):
            yield p
    inp = project_root / "input"
    if not inp.is_dir():
        return
    for p in _glob_workbooks(inp):
        yield p
    rate_resolved = rate_dir.resolve() if rate_dir else None
    for sub in sorted(inp.iterdir()):
        if not sub.is_dir():
            continue
        if rate_resolved is not None and sub.resolve() == rate_resolved:
            continue
        for p in _glob_workbooks(sub):
            yield p


def default_input_workbook(project_root: Path) -> Path | None:
    for p in iter_default_workbook_candidates(project_root):
        if p.is_file():
            return p
    return None


def default_input_search_hint(project_root: Path) -> str:
    return "\n".join(
        [
            "No .xlsx or .xlsb found. Default search order:",
            "  1) input/Rate card/ (or Rate Card / rate card)",
            "  2) input/",
            "  3) input/<any subfolder>/",
            f"Project root: {project_root}",
            "Fix: add a workbook, or run with  --input  path\\to\\file.xlsb",
            "For .xlsb files:  pip install pyxlsb",
        ]
    )


def _main_cost_section_ordered(
    section: dict, sheet_name: str, region_prefix: str | None = None
) -> dict:
    """Build MainCosts section dict with ``tab_name`` immediately before ``zone_headers``."""
    service_type = _correct_service_type_for_tab(
        section.get("service_type", ""), sheet_name, region_prefix
    )
    return {
        "service_type": service_type,
        "cost_category": section.get("cost_category", ""),
        "weight_unit": section.get("weight_unit", ""),
        "tab_name": sheet_name,
        "zone_headers": section.get("zone_headers") or {},
        "pricing": section.get("pricing") or [],
    }


def _collect_main_costs_from_openpyxl(
    workbook_path: Path, region_prefix: str | None
) -> list[dict]:
    import openpyxl

    main_costs: list[dict] = []
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_matches_region_pricing_tab(sheet_name, region_prefix):
                continue
            if sheet_is_ignored_pricing_tab(sheet_name, region_prefix):
                continue
            ws = wb[sheet_name]
            rows = _sheet_to_rows_openpyxl(ws)
            section = extract_first_main_costs_table(rows)
            if not section:
                print(f"[WARN] No first pricing table parsed on sheet {sheet_name!r} (skipped)")
                continue
            main_costs.append(_main_cost_section_ordered(section, sheet_name, region_prefix))
    finally:
        wb.close()
    return main_costs


def _collect_main_costs_from_pyxlsb(
    workbook_path: Path, region_prefix: str | None
) -> list[dict]:
    try:
        from pyxlsb import open_workbook
    except ImportError as e:
        raise SystemExit(
            "Reading .xlsb requires the pyxlsb package. Install with:\n"
            "  pip install pyxlsb"
        ) from e

    main_costs: list[dict] = []
    with open_workbook(str(workbook_path)) as wb:
        for sheet_name in wb.sheets:
            if not sheet_matches_region_pricing_tab(sheet_name, region_prefix):
                continue
            if sheet_is_ignored_pricing_tab(sheet_name, region_prefix):
                continue
            with wb.get_sheet(sheet_name) as sheet:
                rows = _sheet_to_rows_pyxlsb(sheet)
            section = extract_first_main_costs_table(rows)
            if not section:
                print(f"[WARN] No first pricing table parsed on sheet {sheet_name!r} (skipped)")
                continue
            main_costs.append(_main_cost_section_ordered(section, sheet_name, region_prefix))
    return main_costs


def workbook_to_payload(
    workbook_path: Path,
    *,
    client: str = "Unknown",
    carrier: str | None = None,
    validity_date: str | None = None,
    document_currency: str = "",
) -> dict:
    suffix = workbook_path.suffix.lower()
    sheet_names = _workbook_sheet_names(workbook_path, suffix)
    region_prefix = discover_workbook_region_prefix(sheet_names)
    if not region_prefix:
        print(
            "[WARN] No worksheet name begins with two capitals + space (``XX ``); "
            "MainCosts / CountryZoning / header metadata may be empty"
        )

    if suffix == ".xlsb":
        main_costs = _collect_main_costs_from_pyxlsb(workbook_path, region_prefix)
        src = "Rate Card Excel (.xlsb)"
    elif suffix == ".xlsx":
        main_costs = _collect_main_costs_from_openpyxl(workbook_path, region_prefix)
        src = "Rate Card Excel (.xlsx)"
    else:
        raise ValueError(f"Unsupported workbook type {suffix!r}; expected .xlsx or .xlsb")

    country_zoning, accessorial_costs2 = _country_zoning_and_accessorials2_for_suffix(
        workbook_path, suffix, region_prefix
    )

    auto_carrier, auto_currency = _extract_workbook_header_carrier_and_currency(
        workbook_path, suffix, region_prefix
    )
    auto_client = _extract_workbook_client_name(workbook_path, suffix, region_prefix)
    carrier_val = (carrier or "").strip()
    if not carrier_val:
        carrier_val = auto_carrier
    doc_curr = (document_currency or "").strip()
    if not doc_curr:
        doc_curr = auto_currency
    client_val = (client or "").strip()
    if not client_val or client_val == "Unknown":
        client_val = (auto_client or "").strip() or "Unknown"

    total_rows = sum(len(s.get("pricing", [])) for s in main_costs)
    meta = {
        "client": client_val,
        "carrier": carrier_val,
        "validity_date": validity_date or "",
        "document_currency": doc_curr,
        "extraction_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "extraction_source": src,
        "FileName": workbook_path.name,
    }

    stats = {
        "MainCosts_sections": len(main_costs),
        "MainCosts_rows": total_rows,
        "AddedRates_rows": 0,
        "AdditionalCostsPart1_rows": 0,
        "CountryZoning_rows": 0,
        "AdditionalZoning_rows": 0,
        "ZoningMatrix_rows": 0,
        "AdditionalCostsPart2_rows": 0,
        "GoGreenPlusCost_rows": 0,
        "DemandCosts_rows": 0,
        "DemandSurcharge_rows": 0,
        "DemandSurchargeCountries_rows": 0,
    }

    stats["CountryZoning_rows"] = len(country_zoning)
    stats["AccessorialCosts2_rows"] = len(accessorial_costs2)

    return {
        "metadata": meta,
        "MainCosts": main_costs,
        "AddedRates": [],
        "AdditionalCostsPart1": [],
        "AccessorialCosts2": accessorial_costs2,
        "CountryZoning": country_zoning,
        "AdditionalZoning": [],
        "ZoningMatrix": [],
        "AdditionalCostsPart2": [],
        "GoGreenPlusCost": [],
        "DemandCosts": [],
        "DemandSurcharge": [],
        "DemandSurchargeCountries": [],
        "statistics": stats,
    }


def parse_args() -> argparse.Namespace:
    root = Path(__file__).resolve().parent
    default_in = default_input_workbook(root)
    p = argparse.ArgumentParser(
        description="Convert Rate Card Excel: MainCosts ({XX} pricing tabs), CountryZoning ({XX}+ZONE first tab), AccessorialCosts2 (Accessorials tab)."
    )
    p.add_argument(
        "--input",
        type=Path,
        default=default_in,
        help="Path to .xlsx or .xlsb (default: first workbook under input/Rate card, then input/)",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=root / "processing" / "from_rate_card_excel.json",
        help="Output JSON path",
    )
    p.add_argument(
        "--client",
        default="Unknown",
        help="Metadata client (default: first non-empty cell in first 5 rows of first pricing tab for the workbook region prefix)",
    )
    p.add_argument(
        "--carrier",
        default="",
        help="Metadata carrier (default: Country value from sheet, e.g. Sweden)",
    )
    p.add_argument("--validity-date", default="", dest="validity_date")
    p.add_argument(
        "--currency",
        default="",
        dest="document_currency",
        help="Document currency (default: Rate Chart Currency from sheet, e.g. SEK)",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(__file__).resolve().parent
    if not args.input:
        raise SystemExit(default_input_search_hint(root))
    if not args.input.is_file():
        raise SystemExit(f"Input not found: {args.input}")
    suf = args.input.suffix.lower()
    if suf not in WORKBOOK_SUFFIXES:
        raise SystemExit(
            f"Unsupported file type {args.input.suffix!r}. Use .xlsx or .xlsb (got {args.input})."
        )

    payload = workbook_to_payload(
        args.input,
        client=args.client,
        carrier=args.carrier or None,
        validity_date=args.validity_date or None,
        document_currency=args.document_currency,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    st = payload["statistics"]
    print(f"[OK] Wrote {args.output}")
    print(
        f"    MainCosts sections: {st['MainCosts_sections']}, pricing rows: {st['MainCosts_rows']}; "
        f"CountryZoning rows: {st['CountryZoning_rows']}; AccessorialCosts2 rows: {st['AccessorialCosts2_rows']}"
    )


if __name__ == "__main__":
    main()
